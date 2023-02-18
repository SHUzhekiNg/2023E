from openpyxl import load_workbook

# Areas to analyze
geo_names = ['Massachusetts',
             'New York',
             'New Jersey',
             'California',
             'Pennsylvania',
             'Washington',
             'Nevada',
             'Mississippi',
             'New Mexico']

# GDP items to be analyzed
gdp_types = ['All industry total',
             'Construction',
             'Transportation and warehousing',
             'Broadcasting (except Internet) and telecommunications',
             'Finance and insurance',
             'Health care and social assistance',
             'Arts, entertainment, and recreation']

def read_gdp(filename: str = 'gdp.xlsx') -> dict:
    '''Read GDP data from xlsx file'''
    ret = {}
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]

    for geo in geo_names:
        ret[geo] = {}
        for row in sheet.rows:
            if row[1].value == geo:
                desc = str(row[3].value).strip()
                num = float(row[4].value) if not row[4].value == '' else 0
                if desc in gdp_types:
                    ret[geo][desc + '(GDP)'] = num
    return ret

def read_population(filename: str = 'population.xlsx') -> dict:
    '''Read population data from xlsx file'''
    ret = {}
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    for geo in geo_names:
        ret[geo] = {}
        for row in sheet.rows:
            if row[0].value == geo:
                ret[geo]['Population'] = int(row[1].value)
                ret[geo]['Population Density'] = float(row[2].value)
                # ret[geo]['Density Rank'] = int(row[3].value)
    return ret

def read_limiting_magnitude(filename: str = 'limiting_magnitude.xlsx') -> dict:
    '''Reading limiting magnitude from xlsx'''
    ret = {}
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    for geo in geo_names:
        ret[geo] = {}
        for row in sheet.rows:
            if row[0].value == geo:
                ret[geo]['Limiting Magnitude'] = float(row[1].value)
    return ret

def read_last_bus(filename: str = 'last_bus.xlsx') -> dict:
    '''Reading limiting magnitude from xlsx'''
    ret = {}
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    for geo in geo_names:
        ret[geo] = {}
        for row in sheet.rows:
            if row[0].value == geo:
                ret[geo]['Last Bus'] = float(row[1].value)
    return ret

def read_power_consumption(filename: str = 'power_comsumption.xlsx') -> dict:
    '''Read power consumption data'''
    ret = {}
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    for geo in geo_names:
        ret[geo] = {}
        for row in sheet.rows:
            if row[0].value == geo:
                ret[geo]['Power Consumption per Capita per Month'] = float(row[1].value)
    return ret

def read_annual_precipitation(filename: str = 'annual_precipitation.xlsx') -> dict:
    '''Read power consumption data'''
    ret = {}
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    for geo in geo_names:
        ret[geo] = {}
        for row in sheet.rows:
            if row[0].value == geo:
                ret[geo]['Annual Precipitation(in millimetre)'] = float(row[2].value)
    return ret

def read_work_hours(filename: str = 'work_hours.xlsx') -> dict:
    '''Read power consumption data'''
    ret = {}
    workbook = load_workbook(filename)
    sheet = workbook.worksheets[0]
    for geo in geo_names:
        ret[geo] = {}
        for row in sheet.rows:
            if row[0].value == geo:
                ret[geo]['Work hours per Week'] = float(row[1].value)
    return ret

def read_all_data() -> dict:
    '''Read and combine all data'''
    all_data = [read_gdp('../data/GDP by State(tables only).xlsx'),
                read_population('../data/population-density-data-table.xlsx'),
                read_limiting_magnitude('../data/极限星等.xlsx'),
                read_last_bus('../data/公交末班车.xlsx'),
                read_power_consumption('../data/每月人均用电量.xlsx'),
                read_annual_precipitation('../data/年降水量.xlsx'),
                read_work_hours('../data/人均工作时长.xlsx')]
    ret = {}
    for item in all_data:
        for (geo, data) in item.items():
            if not geo in ret.keys():
                ret[geo] = {}
            ret[geo].update(data)
    return ret

def generate_csv() -> str:
    '''Generate csv content'''
    all_data = read_all_data()

    # generate title
    ret = 'Area'
    for k, v in all_data[geo_names[0]].items():
        ret += ', ' + k.replace(',', '')
    ret += '\n'

    # write data
    for geo, data in all_data.items():
        ret += geo
        for k, v in data.items():
            ret += ', ' + str(v)
        ret += '\n'
    return ret

if __name__ == '__main__':
    with open('../data/combined.csv', 'w', encoding='utf-8') as f:
        f.write(generate_csv())

